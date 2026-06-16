import json
import os
import uuid

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

os.environ.setdefault("APP_PASSWORD", "test-password-123")
os.environ.setdefault("SESSION_SECRET_KEY", "test-secret-key-at-least-32-chars!!")
os.environ.setdefault("CLAUDE_API_KEY", "sk-ant-test-key")

from backend.app.services import analysis_service
from backend.app.services.analysis_service import (
    CourseInfo,
    compose_message,
    iter_analysis_events,
    match_applicants,
    parse_csv,
)


def _course(front="앞", back="뒤", desc="설명") -> CourseInfo:
    return CourseInfo(name="AIO1", description=desc, front_msg=front, back_msg=back)


def _fake_analyze(applicant, interview, course):
    return {
        "job_understanding": "높음",
        "course_confidence": "중간",
        "decision_state": "고민 중",
        "real_constraint": "없음",
        "churn_reason": "비용 부담",
        "message_body": f"{applicant['이름']}님 본문",
    }


# --- 순수 로직 ---


def test_parse_csv_strips_and_fills():
    df = parse_csv("이름,연락처\n홍길동,010\n,".encode("utf-8"))
    assert list(df.columns) == ["이름", "연락처"]
    assert df.iloc[1]["이름"] == ""


def test_match_applicants_by_name():
    apps = parse_csv("이름,기수\n홍길동,1기\n김철수,1기".encode("utf-8"))
    ints = parse_csv("이름,점수\n홍길동,5".encode("utf-8"))
    matched, unmatched = match_applicants(apps, ints)
    assert len(matched) == 1
    assert matched[0][0]["이름"] == "홍길동"
    assert len(unmatched) == 1
    assert unmatched[0]["이름"] == "김철수"


def test_compose_message_four_parts():
    msg = compose_message(_course(front="안녕", back="감사"), "본문")
    parts = msg.split("\n\n")
    assert parts[0] == "안녕"
    assert parts[1] == "본문"
    assert parts[2] == "감사"
    assert "http" in parts[3]  # HRD 링크


def test_compose_message_skips_empty_fixed():
    msg = compose_message(_course(front="", back=""), "본문")
    parts = msg.split("\n\n")
    assert parts[0] == "본문"
    assert "http" in parts[1]


def test_iter_analysis_events_progress_and_done():
    apps = "이름,연락처,기수\n홍길동,010,1기\n김철수,011,1기".encode("utf-8")
    ints = "이름,점수\n홍길동,5\n김철수,4".encode("utf-8")
    events = list(iter_analysis_events(apps, ints, _course(), _fake_analyze))

    progress = [e for e in events if e["type"] == "progress"]
    done = [e for e in events if e["type"] == "done"]
    assert len(progress) == 2
    assert progress[-1] == {"type": "progress", "current": 2, "total": 2}
    assert len(done) == 1
    rows = done[0]["rows"]
    assert len(rows) == 2
    assert rows[0]["name"] == "홍길동"
    assert rows[0]["failed"] is False
    assert rows[0]["message"].startswith("앞")
    assert "홍길동님 본문" in rows[0]["message"]


def test_iter_analysis_events_unmatched_marked_failed():
    apps = "이름,연락처\n홍길동,010\n매칭없음,011".encode("utf-8")
    ints = "이름,점수\n홍길동,5".encode("utf-8")
    events = list(iter_analysis_events(apps, ints, _course(), _fake_analyze))
    rows = [e for e in events if e["type"] == "done"][0]["rows"]
    by_name = {r["name"]: r for r in rows}
    assert by_name["홍길동"]["failed"] is False
    assert by_name["매칭없음"]["failed"] is True
    assert by_name["매칭없음"]["churn_reason"] == "분석 실패"


def test_iter_analysis_events_empty_csv_total_zero():
    """헤더만 있는 빈 CSV는 progress 없이 done(rows=[])만 낸다 (total=0 경계)"""
    apps = "이름,연락처".encode("utf-8")  # 데이터 행 0개
    ints = "이름,점수".encode("utf-8")
    events = list(iter_analysis_events(apps, ints, _course(), _fake_analyze))
    assert [e["type"] for e in events] == ["done"]
    assert events[0]["rows"] == []


def test_iter_analysis_events_parse_error_yields_error():
    """깨진/빈 바이트로 파싱 실패 시 error 이벤트를 낸다 (raw 예외 비노출)"""
    events = list(iter_analysis_events(b"", b"", _course(), _fake_analyze))
    assert events == [{"type": "error", "message": "CSV 형식을 확인해주세요"}]


def test_iter_analysis_events_too_many_applicants(monkeypatch):
    """MAX_APPLICANTS 초과 시 분석을 시작하지 않고 error 이벤트를 낸다"""
    monkeypatch.setattr(analysis_service, "MAX_APPLICANTS", 1)
    apps = "이름\n홍길동\n김철수".encode("utf-8")
    ints = "이름\n홍길동\n김철수".encode("utf-8")
    events = list(iter_analysis_events(apps, ints, _course(), _fake_analyze))
    assert events[0]["type"] == "error"
    assert "너무 많습니다" in events[0]["message"]


def test_build_excel_empty_rows_header_only():
    """0행이어도 헤더만 있는 유효한 엑셀을 만든다"""
    import io as _io

    from openpyxl import load_workbook

    from backend.app.services.analysis_service import build_excel

    wb = load_workbook(_io.BytesIO(build_excel([])))
    ws = wb.active
    assert ws.max_row == 1  # 헤더 행만
    assert ws[1][0].value == "기수"


def test_iter_analysis_events_failure_continues():
    def flaky(applicant, interview, course):
        if applicant["이름"] == "실패자":
            raise RuntimeError("Claude API 오류")
        return _fake_analyze(applicant, interview, course)

    apps = "이름,연락처\n실패자,010\n정상,011".encode("utf-8")
    ints = "이름,점수\n실패자,5\n정상,4".encode("utf-8")
    events = list(iter_analysis_events(apps, ints, _course(), flaky))
    rows = [e for e in events if e["type"] == "done"][0]["rows"]
    by_name = {r["name"]: r for r in rows}
    assert by_name["실패자"]["failed"] is True
    assert by_name["정상"]["failed"] is False  # 실패해도 나머지 계속


# --- 엔드포인트 ---


@pytest.fixture
def test_engine():
    from backend.app.database import Base
    from backend.app.models import course as _  # noqa: F401

    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    return engine


@pytest.fixture
def client(test_engine, monkeypatch):
    from fastapi.testclient import TestClient

    from backend.app.database import get_db
    from backend.app.main import app

    TestSession = sessionmaker(bind=test_engine, autocommit=False, autoflush=False)

    def override_get_db():
        db = TestSession()
        try:
            yield db
        finally:
            db.close()

    # 실제 Claude 호출 차단 — 가짜 analyzer 주입
    monkeypatch.setattr(analysis_service, "make_claude_analyzer", lambda: _fake_analyze)

    app.dependency_overrides[get_db] = override_get_db
    test_client = TestClient(app)
    yield test_client
    app.dependency_overrides.clear()


@pytest.fixture
def auth_client(client):
    client.post("/api/v1/auth/login", json={"password": "test-password-123"})
    return client


@pytest.fixture
def seed_course(test_engine):
    from backend.app.models.course import Course

    TestSession = sessionmaker(bind=test_engine)

    def _seed(name="AIO1", front_msg="앞", back_msg="뒤") -> str:
        session = TestSession()
        course = Course(id=str(uuid.uuid4()), name=name, front_msg=front_msg, back_msg=back_msg)
        session.add(course)
        session.commit()
        cid = course.id
        session.close()
        return cid

    return _seed


def _files():
    return {
        "applicant_csv": ("a.csv", "이름,연락처\n홍길동,010".encode("utf-8"), "text/csv"),
        "interview_csv": ("b.csv", "이름,점수\n홍길동,5".encode("utf-8"), "text/csv"),
    }


def test_start_requires_auth(client, seed_course):
    cid = seed_course()
    res = client.post("/api/v1/analysis/start", data={"course_id": cid}, files=_files())
    assert res.status_code == 401


def test_start_course_not_found(auth_client):
    res = auth_client.post(
        "/api/v1/analysis/start", data={"course_id": "nope"}, files=_files()
    )
    assert res.status_code == 404


def test_start_rejects_oversized_upload(auth_client, seed_course, monkeypatch):
    """업로드 CSV가 상한을 넘으면 413"""
    from backend.app.routers import analysis as analysis_router

    monkeypatch.setattr(analysis_router, "MAX_UPLOAD_BYTES", 10)
    cid = seed_course()
    big = ("이름,연락처\n" + "홍길동,010\n" * 100).encode("utf-8")
    res = auth_client.post(
        "/api/v1/analysis/start",
        data={"course_id": cid},
        files={
            "applicant_csv": ("a.csv", big, "text/csv"),
            "interview_csv": ("b.csv", big, "text/csv"),
        },
    )
    assert res.status_code == 413


def test_start_streams_progress_and_done(auth_client, seed_course):
    cid = seed_course()
    res = auth_client.post("/api/v1/analysis/start", data={"course_id": cid}, files=_files())
    assert res.status_code == 200
    assert res.headers["content-type"].startswith("text/event-stream")

    events = [
        json.loads(line[len("data: ") :])
        for line in res.text.splitlines()
        if line.startswith("data: ")
    ]
    assert any(e["type"] == "progress" for e in events)
    done = [e for e in events if e["type"] == "done"]
    assert len(done) == 1
    assert done[0]["rows"][0]["name"] == "홍길동"


# --- 엑셀 (Story 3.3) ---


def _result_row(name="홍길동", failed=False):
    return {
        "cohort": "1기",
        "name": name,
        "phone": "010",
        "job_understanding": "높음",
        "course_confidence": "중간",
        "decision_state": "고민",
        "real_constraint": "없음",
        "churn_reason": "분석 실패" if failed else "비용 부담",
        "message": "" if failed else "안녕하세요",
        "failed": failed,
    }


def test_build_excel_headers_and_values():
    import io as _io

    from openpyxl import load_workbook

    from backend.app.services.analysis_service import build_excel

    content = build_excel([_result_row()])
    wb = load_workbook(_io.BytesIO(content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers[0] == "기수"
    assert "권장 독려문자멘트" in headers
    assert ws[2][1].value == "홍길동"


def test_build_excel_failed_row_red():
    import io as _io

    from openpyxl import load_workbook

    from backend.app.services.analysis_service import build_excel

    content = build_excel([_result_row(failed=True)])
    wb = load_workbook(_io.BytesIO(content))
    ws = wb.active
    # 2행(첫 데이터 행)이 빨간 채움이어야 한다
    assert ws[2][0].fill.start_color.rgb.endswith("FFC7CE")


def test_excel_requires_auth(client):
    res = client.post(
        "/api/v1/analysis/excel",
        json={"course_name": "AIO1", "analyzed_at": "", "rows": [_result_row()]},
    )
    assert res.status_code == 401


def test_excel_downloads_with_filename(auth_client):
    res = auth_client.post(
        "/api/v1/analysis/excel",
        json={"course_name": "AIO1", "analyzed_at": "2026-06-17T09:23:00", "rows": [_result_row()]},
    )
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]
    disposition = res.headers["content-disposition"]
    assert "filename*=UTF-8''" in disposition
    assert "20260617" in disposition  # 날짜 반영
    assert len(res.content) > 0
