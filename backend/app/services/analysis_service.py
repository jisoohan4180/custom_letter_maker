"""CSV 분석 비즈니스 로직 — 매칭, Claude 분석, 멘트 4단 조합.

개인정보(이름·연락처)는 분석 중에만 메모리에 존재하고 DB에 저장하지 않는다 (NFR3).
SSE 진행률 이벤트는 sync generator(iter_analysis_events)로 만들고, 라우터가 async로 감싼다.
"""
import io
import os
from dataclasses import dataclass
from typing import Callable, Iterator

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# 이름 컬럼을 매칭 키로 사용 (TRD: 이름 기준 자동 매칭)
NAME_COLUMN = "이름"
PHONE_COLUMN = "연락처"
COHORT_COLUMN = "기수"

CLAUDE_MODEL = os.getenv("CLAUDE_MODEL", "claude-haiku-4-5")
HRD_REGISTER_URL = os.getenv("HRD_REGISTER_URL", "https://hrd.example.com/register")
# 외부 API 호출 timeout (초) — 한 명이 hang 돼도 다음으로 빨리 넘어가도록
CLAUDE_TIMEOUT = float(os.getenv("CLAUDE_TIMEOUT", "30"))
# 한 번에 분석할 지원자 수 상한 (무바운드 순차 호출/비용 폭증 방지)
MAX_APPLICANTS = int(os.getenv("MAX_APPLICANTS", "500"))

# Claude 강제 도구 호출용 스키마 — 4항목 분석 + 이탈사유 + 멘트 본문
ANALYSIS_TOOL = {
    "name": "record_analysis",
    "description": "지원자의 이탈 사유 4항목 분석 결과와 개인화 독려 멘트 본문을 기록한다.",
    "input_schema": {
        "type": "object",
        "properties": {
            "job_understanding": {"type": "string", "description": "AI/직무이해도 분석 한 줄"},
            "course_confidence": {"type": "string", "description": "과정확신도 분석 한 줄"},
            "decision_state": {"type": "string", "description": "의사결정상태 분석 한 줄"},
            "real_constraint": {"type": "string", "description": "현실제약 여부 분석 한 줄"},
            "churn_reason": {"type": "string", "description": "이탈 예측 사유 (종합 한 줄)"},
            "message_body": {
                "type": "string",
                "description": "개인화 독려 멘트 본문 (앞/뒤 고정 멘트와 링크는 제외)",
            },
        },
        "required": [
            "job_understanding",
            "course_confidence",
            "decision_state",
            "real_constraint",
            "churn_reason",
            "message_body",
        ],
        "additionalProperties": False,
    },
}


@dataclass
class CourseInfo:
    """분석에 필요한 과정 데이터 스냅샷 (DB 세션과 분리)."""

    name: str
    description: str
    front_msg: str
    back_msg: str


# analyze(applicant: dict, interview: dict, course: CourseInfo) -> dict(fields)
Analyzer = Callable[[dict, dict, CourseInfo], dict]


def parse_csv(content: bytes) -> pd.DataFrame:
    """CSV 바이트를 모든 컬럼 문자열로 읽고 NaN은 빈 문자열로 채운다."""
    df = pd.read_csv(io.BytesIO(content), dtype=str).fillna("")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _cell(row: dict, column: str) -> str:
    return str(row.get(column, "")).strip()


def match_applicants(
    applicants: pd.DataFrame, interviews: pd.DataFrame
) -> tuple[list[tuple[dict, dict]], list[dict]]:
    """이름 기준으로 두 CSV를 매칭한다. (매칭 쌍, 매칭 실패 신청자) 반환."""
    interview_by_name: dict[str, dict] = {}
    for _, row in interviews.iterrows():
        name = _cell(row.to_dict(), NAME_COLUMN)
        if name and name not in interview_by_name:
            interview_by_name[name] = row.to_dict()

    matched: list[tuple[dict, dict]] = []
    unmatched: list[dict] = []
    for _, row in applicants.iterrows():
        applicant = row.to_dict()
        name = _cell(applicant, NAME_COLUMN)
        interview = interview_by_name.get(name)
        if interview is not None:
            matched.append((applicant, interview))
        else:
            unmatched.append(applicant)
    return matched, unmatched


def compose_message(course: CourseInfo, body: str) -> str:
    """앞 고정 멘트 + AI 본문 + 뒤 고정 멘트 + HRD 등록 링크 4단 조합 (빈 부분 생략)."""
    parts = [course.front_msg.strip(), (body or "").strip(), course.back_msg.strip(), HRD_REGISTER_URL]
    return "\n\n".join(part for part in parts if part)


def _row(applicant: dict, course: CourseInfo, fields: dict | None, failed: bool) -> dict:
    return {
        "cohort": _cell(applicant, COHORT_COLUMN),
        "name": _cell(applicant, NAME_COLUMN),
        "phone": _cell(applicant, PHONE_COLUMN),
        "job_understanding": (fields or {}).get("job_understanding", ""),
        "course_confidence": (fields or {}).get("course_confidence", ""),
        "decision_state": (fields or {}).get("decision_state", ""),
        "real_constraint": (fields or {}).get("real_constraint", ""),
        "churn_reason": "분석 실패" if failed else (fields or {}).get("churn_reason", ""),
        "message": "" if failed else compose_message(course, (fields or {}).get("message_body", "")),
        "failed": failed,
    }


def iter_analysis_events(
    applicant_bytes: bytes,
    interview_bytes: bytes,
    course: CourseInfo,
    analyze: Analyzer,
) -> Iterator[dict]:
    """진행률 이벤트를 순차 생성한다.

    각 지원자를 1명씩 분석하고, 실패해도 다음 지원자를 계속 진행한다 (FR17).
    마지막에 전체 결과 rows를 담은 done 이벤트를 낸다.
    """
    try:
        applicants = parse_csv(applicant_bytes)
        interviews = parse_csv(interview_bytes)
    except Exception:
        # 깨진/빈 CSV 등 파싱 실패 — 일반화된 에러 이벤트 (상세는 노출 안 함)
        yield {"type": "error", "message": "CSV 형식을 확인해주세요"}
        return

    matched, unmatched = match_applicants(applicants, interviews)

    total = len(matched) + len(unmatched)
    if total > MAX_APPLICANTS:
        yield {"type": "error", "message": f"지원자가 너무 많습니다 (최대 {MAX_APPLICANTS}명)"}
        return

    rows: list[dict] = []
    done = 0

    for applicant, interview in matched:
        done += 1
        try:
            fields = analyze(applicant, interview, course)
            rows.append(_row(applicant, course, fields, failed=False))
        except Exception:
            # 개별 지원자 분석 실패 → 실패 행으로 표시하고 계속 진행
            rows.append(_row(applicant, course, None, failed=True))
        yield {"type": "progress", "current": done, "total": total}

    # 매칭 실패 신청자는 분석 실패로 분류
    for applicant in unmatched:
        done += 1
        rows.append(_row(applicant, course, None, failed=True))
        yield {"type": "progress", "current": done, "total": total}

    yield {"type": "done", "rows": rows}


# 엑셀 출력 컬럼 (AC 3.3) 과 AnalysisRow 키 매핑
EXCEL_COLUMNS: list[tuple[str, str]] = [
    ("기수", "cohort"),
    ("이름", "name"),
    ("연락처", "phone"),
    ("AI/직무이해도", "job_understanding"),
    ("과정확신도", "course_confidence"),
    ("의사결정상태", "decision_state"),
    ("현실제약", "real_constraint"),
    ("이탈예측사유", "churn_reason"),
    ("권장 독려문자멘트", "message"),
]
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def build_excel(rows: list[dict]) -> bytes:
    """분석 결과를 xlsx 바이트로 만든다. 분석 실패 행은 빨간 배경(PatternFill)."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "독려문자"

    headers = [label for label, _ in EXCEL_COLUMNS]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append([str(row.get(key, "")) for _, key in EXCEL_COLUMNS])
        if row.get("failed"):
            for cell in sheet[sheet.max_row]:
                cell.fill = _RED_FILL

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def make_claude_analyzer() -> Analyzer:
    """실제 Claude API를 호출하는 analyzer를 만든다. (테스트는 이 함수를 monkeypatch)"""
    import anthropic

    # 프로젝트는 CLAUDE_API_KEY 를 사용하므로 명시적으로 주입 (SDK 기본은 ANTHROPIC_API_KEY)
    client = anthropic.Anthropic(api_key=os.environ["CLAUDE_API_KEY"], timeout=CLAUDE_TIMEOUT)

    def analyze(applicant: dict, interview: dict, course: CourseInfo) -> dict:
        applicant_text = "\n".join(f"{k}: {v}" for k, v in applicant.items())
        interview_text = "\n".join(f"{k}: {v}" for k, v in interview.items())
        prompt = (
            "다음은 HRD 교육과정 지원자 정보와 인터뷰 평가입니다. "
            "지원자의 이탈 가능성을 4항목(AI/직무이해도·과정확신도·의사결정상태·현실제약)으로 "
            "분석하고, 등록을 독려하는 개인화 멘트 본문을 작성하세요.\n\n"
            f"[과정 설명]\n{course.description or '(없음)'}\n\n"
            f"[지원자 정보]\n{applicant_text}\n\n"
            f"[인터뷰 평가]\n{interview_text}"
        )
        response = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=1024,
            tools=[ANALYSIS_TOOL],
            tool_choice={"type": "tool", "name": "record_analysis"},
            messages=[{"role": "user", "content": prompt}],
        )
        for block in response.content:
            if block.type == "tool_use" and block.name == "record_analysis":
                return dict(block.input)
        raise ValueError("Claude 분석 결과를 받지 못했습니다")

    return analyze
